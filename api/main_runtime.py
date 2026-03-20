# -*- coding: utf-8 -*-
"""
本地热修复版 Web 入口。
用于绕开共享盘空间不足导致的上传、统计与静态页读写失败。
"""

import json
import logging
import multiprocessing
import re
import shutil
import sys
import tempfile
import threading
import time
import uuid
from datetime import datetime, timedelta
from io import BytesIO
from pathlib import Path

from fastapi import BackgroundTasks, FastAPI, File, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse
from openpyxl import load_workbook
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.errors import RateLimitExceeded
from slowapi.util import get_remote_address


PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)

app = FastAPI(
    title="成都地区地基基础分析系统",
    description="Web API for Foundation Analysis System",
    version="2.1.0",
)
limiter = Limiter(key_func=get_remote_address)
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:8086",
        "http://127.0.0.1:8086",
        "http://localhost:3000",
    ],
    allow_credentials=True,
    allow_methods=["GET", "POST"],
    allow_headers=["*"],
)

RUNTIME_ROOT = Path(tempfile.gettempdir()) / "wenxing-web-runtime"
UPLOAD_DIR = RUNTIME_ROOT / "uploads"
OUTPUT_DIR = RUNTIME_ROOT / "outputs"
STATS_DIR = RUNTIME_ROOT / "stats"
STATS_FILE = STATS_DIR / "usage_stats.json"
for path in (RUNTIME_ROOT, UPLOAD_DIR, OUTPUT_DIR, STATS_DIR):
    path.mkdir(parents=True, exist_ok=True)

MAX_ACTIVE_TASKS = 30
TASK_TTL_SECONDS = 24 * 3600
MAX_FILE_SIZE_MB = 30
ANALYSIS_TIMEOUT_SECONDS = 300
CLEANUP_INTERVAL_SECONDS = 1800
INVALID_FILENAME_CHARS = re.compile(r'[<>:"/\\|?*\x00-\x1f]')
tasks: dict[str, dict] = {}
stats_lock = threading.RLock()


INDEX_HTML = """<!DOCTYPE html>
<html lang="zh-CN">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>地基基础分析报告生成</title>
  <style>
    :root { --primary:#1f6feb; --primary-dark:#1558b0; --bg:#f4f8fc; --card:#fff; --text:#1f2328; --muted:#667085; --border:#d0d7de; --danger:#d92d20; --danger-bg:#fef3f2; --success:#16a34a; }
    * { box-sizing:border-box; }
    body { margin:0; font-family:"Microsoft YaHei","PingFang SC",sans-serif; background:linear-gradient(180deg,#eef6ff 0%,var(--bg) 100%); color:var(--text); }
    .wrap { max-width:880px; margin:0 auto; padding:32px 16px 48px; }
    .card { background:var(--card); border-radius:18px; box-shadow:0 12px 36px rgba(15,23,42,.08); padding:24px; margin-bottom:18px; }
    h1 { margin:0 0 8px; font-size:30px; text-align:center; color:var(--primary); }
    .subtitle { margin:0 0 24px; text-align:center; color:var(--muted); font-size:14px; }
    .toolbar { display:flex; justify-content:space-between; align-items:center; gap:12px; flex-wrap:wrap; padding:16px; border:1px solid #dbeafe; border-radius:14px; background:#f8fbff; margin-bottom:18px; }
    .toolbar a { color:var(--primary); text-decoration:none; font-weight:600; }
    .upload-box { border:2px dashed var(--border); border-radius:16px; padding:28px 20px; text-align:center; background:#fbfdff; cursor:pointer; display:block; }
    .upload-box strong { display:block; margin-bottom:8px; font-size:18px; }
    .upload-box span { color:var(--muted); word-break:break-all; }
    .params { display:none; margin-top:18px; border:1px solid var(--border); border-radius:14px; padding:18px; }
    .question { margin-bottom:18px; }
    .question:last-child { margin-bottom:0; }
    .question-title { font-weight:600; margin-bottom:8px; }
    .question-hint { margin-bottom:10px; color:var(--primary); font-size:13px; }
    .options { display:flex; gap:18px; flex-wrap:wrap; }
    .actions { display:flex; gap:12px; flex-wrap:wrap; margin-top:18px; }
    button { border:none; border-radius:12px; padding:14px 20px; font-size:15px; cursor:pointer; }
    .primary-btn { background:linear-gradient(135deg,var(--primary) 0%,var(--primary-dark) 100%); color:#fff; font-weight:600; flex:1 1 220px; }
    .secondary-btn { background:#fff; border:1px solid var(--border); color:var(--text); }
    .progress { display:none; margin-top:18px; }
    .bar { height:10px; border-radius:999px; background:#e5e7eb; overflow:hidden; }
    .bar > div { width:0%; height:100%; background:linear-gradient(90deg,var(--primary) 0%,#22c55e 100%); transition:width .3s ease; }
    .progress-text { margin-top:10px; text-align:center; color:var(--muted); }
    .error { display:none; margin-top:18px; padding:14px 16px; background:var(--danger-bg); border:1px solid #fecdca; border-radius:12px; color:var(--danger); }
    .result { display:none; margin-top:18px; padding:18px; border-radius:14px; background:#effdf3; border:1px solid #b7ebc6; }
    .stats { display:grid; grid-template-columns:repeat(auto-fit, minmax(180px, 1fr)); gap:14px; }
    .stat-item { padding:18px; border-radius:14px; background:#f8fbff; border:1px solid #e5eefb; }
    .stat-item strong { display:block; font-size:28px; color:var(--primary); margin-top:8px; }
    .history { margin-top:18px; color:var(--muted); font-size:14px; line-height:1.8; }
    @media (max-width:640px) { h1 { font-size:24px; } .card { padding:18px; } }
  </style>
</head>
<body>
  <div class="wrap">
    <div class="card">
      <h1>地基基础分析报告生成</h1>
      <p class="subtitle">上传 Excel 文件，自动生成分析报告</p>
      <div class="toolbar">
        <div>首次使用请先下载模板，按模板填写后再上传。</div>
        <a href="/api/template">下载模板</a>
      </div>
      <label class="upload-box" for="fileInput">
        <strong id="uploadTitle">选择 Excel 文件</strong>
        <span id="uploadHint">支持 .xlsx / .xls</span>
      </label>
      <input id="fileInput" type="file" accept=".xlsx,.xls" hidden>
      <div class="params" id="paramsSection">
        <div class="question">
          <div class="question-title">1. 持力层位于地下水位以上吗？</div>
          <div class="question-hint">是：γ=20；否：γ=10</div>
          <div class="options">
            <label><input type="radio" name="waterLevel" value="yes" checked> 是</label>
            <label><input type="radio" name="waterLevel" value="no"> 否</label>
          </div>
        </div>
        <div class="question" id="question2" style="display:none;">
          <div class="question-title">2. 粉土黏粒含量是否 ≥ 10%？</div>
          <div class="options">
            <label><input type="radio" name="siltClay" value="yes" checked> 是</label>
            <label><input type="radio" name="siltClay" value="no"> 否</label>
          </div>
        </div>
        <div class="question" id="question3" style="display:none;">
          <div class="question-title">3. 粉质黏土 e、Il 是否 ≥ 0.85？</div>
          <div class="options">
            <label><input type="radio" name="siltyClay" value="yes" checked> 是</label>
            <label><input type="radio" name="siltyClay" value="no"> 否</label>
          </div>
        </div>
      </div>
      <div class="actions">
        <button class="primary-btn" id="runButton">开始分析</button>
        <button class="secondary-btn" id="resetButton" type="button">重新选择</button>
      </div>
      <div class="progress" id="progressSection">
        <div class="bar"><div id="progressBar"></div></div>
        <div class="progress-text" id="progressText">准备中...</div>
      </div>
      <div class="error" id="errorBox"></div>
      <div class="result" id="resultBox">
        <div style="font-weight:600; margin-bottom:10px; color:var(--success);">分析完成</div>
        <div style="margin-bottom:14px;">报告已生成，可以直接下载。</div>
        <button class="primary-btn" id="downloadButton" type="button">下载分析报告</button>
      </div>
    </div>
    <div class="card">
      <div class="stats">
        <div class="stat-item">今日分析次数<strong id="todayCount">-</strong></div>
        <div class="stat-item">最近 7 天分析次数<strong id="weekCount">-</strong></div>
      </div>
      <div class="history" id="historyBox">统计加载中...</div>
    </div>
  </div>
  <script>
    let selectedFile = null;
    let currentTaskId = null;
    let pollTimer = null;
    const fileInput = document.getElementById('fileInput');
    const uploadTitle = document.getElementById('uploadTitle');
    const uploadHint = document.getElementById('uploadHint');
    const paramsSection = document.getElementById('paramsSection');
    const question2 = document.getElementById('question2');
    const question3 = document.getElementById('question3');
    const runButton = document.getElementById('runButton');
    const resetButton = document.getElementById('resetButton');
    const progressSection = document.getElementById('progressSection');
    const progressBar = document.getElementById('progressBar');
    const progressText = document.getElementById('progressText');
    const errorBox = document.getElementById('errorBox');
    const resultBox = document.getElementById('resultBox');
    const downloadButton = document.getElementById('downloadButton');

    async function parseResponsePayload(response) {
      const text = await response.text();
      if (!text) return {};
      const contentType = response.headers.get('content-type') || '';
      if (contentType.includes('application/json')) {
        try { return JSON.parse(text); } catch (error) { console.error('JSON 解析失败:', error, text); }
      }
      return { detail: text, message: text, raw: text };
    }

    async function fetchJson(url, options) {
      const response = await fetch(url, options);
      const payload = await parseResponsePayload(response);
      if (!response.ok) throw new Error(payload.detail || payload.message || ('请求失败 (' + response.status + ')'));
      return payload;
    }

    function showError(message) {
      errorBox.textContent = message;
      errorBox.style.display = 'block';
    }

    function hideError() {
      errorBox.style.display = 'none';
      errorBox.textContent = '';
    }

    function resetPage() {
      selectedFile = null;
      currentTaskId = null;
      if (pollTimer) clearInterval(pollTimer);
      fileInput.value = '';
      uploadTitle.textContent = '选择 Excel 文件';
      uploadHint.textContent = '支持 .xlsx / .xls';
      paramsSection.style.display = 'none';
      question2.style.display = 'none';
      question3.style.display = 'none';
      progressSection.style.display = 'none';
      resultBox.style.display = 'none';
      progressBar.style.width = '0%';
      progressText.textContent = '准备中...';
      runButton.disabled = false;
      runButton.textContent = '开始分析';
      hideError();
    }

    async function handleFileSelect(file) {
      if (!file || !/\\.(xlsx|xls)$/i.test(file.name)) {
        showError('请选择 Excel 文件（.xlsx 或 .xls）');
        return;
      }
      selectedFile = file;
      uploadTitle.textContent = '已选择文件';
      uploadHint.textContent = file.name;
      resultBox.style.display = 'none';
      hideError();
      const formData = new FormData();
      formData.append('file', file);
      try {
        const data = await fetchJson('/api/precheck', { method: 'POST', body: formData });
        paramsSection.style.display = 'block';
        question2.style.display = data.has_silt ? 'block' : 'none';
        question3.style.display = data.has_silty_clay ? 'block' : 'none';
      } catch (error) {
        paramsSection.style.display = 'block';
        question2.style.display = 'block';
        question3.style.display = 'block';
      }
    }

    async function startAnalysis() {
      if (!selectedFile) {
        showError('请先选择 Excel 文件');
        return;
      }
      hideError();
      resultBox.style.display = 'none';
      runButton.disabled = true;
      runButton.textContent = '分析中...';
      progressSection.style.display = 'block';
      progressBar.style.width = '10%';
      progressText.textContent = '正在上传文件...';
      const waterLevel = document.querySelector('input[name="waterLevel"]:checked')?.value || 'yes';
      const siltClay = document.querySelector('input[name="siltClay"]:checked')?.value || 'yes';
      const siltyClay = document.querySelector('input[name="siltyClay"]:checked')?.value || 'yes';
      const formData = new FormData();
      formData.append('file', selectedFile);
      formData.append('water_level_above', waterLevel === 'yes' ? 'true' : 'false');
      formData.append('silt_clay_content_ge_10', siltClay === 'yes' ? 'true' : 'false');
      formData.append('silty_clay_e_il_ge_085', siltyClay === 'yes' ? 'true' : 'false');
      try {
        const data = await fetchJson('/api/upload', { method: 'POST', body: formData });
        currentTaskId = data.task_id;
        progressText.textContent = '任务已创建，正在分析...';
        startPolling();
      } catch (error) {
        showError(error.message);
        runButton.disabled = false;
        runButton.textContent = '开始分析';
        progressSection.style.display = 'none';
      }
    }

    function startPolling() {
      let progress = 15;
      pollTimer = setInterval(async () => {
        try {
          const data = await fetchJson('/api/status/' + currentTaskId);
          progress = Math.min(progress + Math.random() * 12, 92);
          progressBar.style.width = progress.toFixed(0) + '%';
          progressText.textContent = '正在分析... ' + progress.toFixed(0) + '%';
          if (data.status === 'completed') {
            clearInterval(pollTimer);
            progressBar.style.width = '100%';
            progressText.textContent = '分析完成';
            runButton.disabled = false;
            runButton.textContent = '开始分析';
            resultBox.style.display = 'block';
          } else if (data.status === 'failed') {
            clearInterval(pollTimer);
            runButton.disabled = false;
            runButton.textContent = '开始分析';
            showError(data.message || '分析失败');
          }
        } catch (error) {
          clearInterval(pollTimer);
          runButton.disabled = false;
          runButton.textContent = '开始分析';
          showError(error.message || '无法获取任务状态');
        }
      }, 1000);
    }

    async function loadStats() {
      try {
        const today = await fetchJson('/api/stats/today');
        const summary = await fetchJson('/api/stats/summary?days=7');
        document.getElementById('todayCount').textContent = today.analysis || 0;
        document.getElementById('weekCount').textContent = summary.period_analysis || 0;
        const history = (summary.daily || []).map(item => item.date + '：' + (item.analysis || 0) + ' 次').join('<br>');
        document.getElementById('historyBox').innerHTML = history || '暂无统计数据';
      } catch (error) {
        document.getElementById('historyBox').textContent = '统计加载失败';
      }
    }

    fileInput.addEventListener('change', event => { if (event.target.files.length > 0) handleFileSelect(event.target.files[0]); });
    runButton.addEventListener('click', startAnalysis);
    resetButton.addEventListener('click', resetPage);
    downloadButton.addEventListener('click', () => { if (currentTaskId) window.location.href = '/api/download/' + currentTaskId; });
    loadStats();
  </script>
</body>
</html>
"""


def is_allowed_excel_file(filename: str) -> bool:
    return bool(filename) and Path(filename).suffix.lower() in {".xlsx", ".xls"}


def sanitize_upload_filename(filename: str) -> str:
    original_name = Path(filename or "").name
    suffix = Path(original_name).suffix.lower()
    stem = Path(original_name).stem
    safe_stem = INVALID_FILENAME_CHARS.sub("_", stem)
    safe_stem = re.sub(r"\s+", "_", safe_stem).strip(" ._")
    if not safe_stem:
        safe_stem = "upload"
    safe_suffix = INVALID_FILENAME_CHARS.sub("", suffix)[:10]
    return f"{safe_stem[:80]}{safe_suffix}"


def _load_stats() -> dict:
    if STATS_FILE.exists():
        try:
            return json.loads(STATS_FILE.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            logger.exception("加载统计数据失败")
    return {"daily": {}, "total": 0}


def _save_stats(stats: dict) -> None:
    STATS_DIR.mkdir(parents=True, exist_ok=True)
    STATS_FILE.write_text(json.dumps(stats, ensure_ascii=False, indent=2), encoding="utf-8")


def record_usage(event_type: str = "analysis") -> None:
    with stats_lock:
        stats = _load_stats()
        today = datetime.now().strftime("%Y-%m-%d")
        stats["daily"].setdefault(today, {"analysis": 0, "download": 0, "precheck": 0})
        stats["daily"][today].setdefault(event_type, 0)
        stats["daily"][today][event_type] += 1
        stats["total"] = stats.get("total", 0) + 1
        _save_stats(stats)


def get_today_stats() -> dict:
    with stats_lock:
      stats = _load_stats()
      today = datetime.now().strftime("%Y-%m-%d")
      daily = stats.get("daily", {}).get(today, {"analysis": 0, "download": 0, "precheck": 0})
      return {
          "date": today,
          "analysis": daily.get("analysis", 0),
          "download": daily.get("download", 0),
          "precheck": daily.get("precheck", 0),
          "total": sum(daily.values()),
      }


def get_stats_summary(days: int = 7) -> dict:
    with stats_lock:
        stats = _load_stats()
        daily_stats = []
        for offset in range(days - 1, -1, -1):
            date = (datetime.now() - timedelta(days=offset)).strftime("%Y-%m-%d")
            daily = stats.get("daily", {}).get(date, {"analysis": 0, "download": 0, "precheck": 0})
            daily_stats.append(
                {
                    "date": date,
                    "analysis": daily.get("analysis", 0),
                    "download": daily.get("download", 0),
                    "precheck": daily.get("precheck", 0),
                    "total": sum(daily.values()),
                }
            )
        return {
            "period_days": days,
            "period_total": sum(item["total"] for item in daily_stats),
            "period_analysis": sum(item["analysis"] for item in daily_stats),
            "all_time_total": stats.get("total", 0),
            "daily": daily_stats,
            "today": get_today_stats(),
        }


def cleanup_files(task: dict) -> None:
    for key in ("input_file", "output_path", "output_file"):
        value = task.get(key)
        if not value:
            continue
        path = Path(value)
        if path.exists():
            try:
                path.unlink()
            except OSError:
                logger.warning("清理文件失败: %s", path)


def cleanup_old_tasks() -> None:
    now = datetime.now()
    expired = []
    for task_id, task in list(tasks.items()):
        created_at = task.get("created_at")
        if not created_at:
            continue
        try:
            created_time = datetime.fromisoformat(created_at)
        except ValueError:
            continue
        if (now - created_time).total_seconds() > TASK_TTL_SECONDS:
            expired.append(task_id)
    for task_id in expired:
        task = tasks.pop(task_id, None)
        if task:
            cleanup_files(task)


def scheduled_cleanup() -> None:
    while True:
        try:
            cleanup_old_tasks()
        except Exception:
            logger.exception("定时清理失败")
        time.sleep(CLEANUP_INTERVAL_SECONDS)


def _run_analysis_in_process(input_path: str, output_path: str, params: dict, queue: multiprocessing.Queue) -> None:
    try:
        from wenxing2 import run_analysis

        def ask_yes_no(title, message):
            if "地下水位" in message or "持力层位置" in title:
                return params.get("water_level_above", True)
            if "粉土" in title and "黏粒含量" in message:
                return params.get("silt_clay_content_ge_10", True)
            if "粉质黏土" in title or "粉质粘土" in title:
                return params.get("silty_clay_e_il_ge_085", True)
            return True

        run_analysis(None, input_path, output_path, ask_yes_no)
        queue.put({"success": True, "error": None})
    except Exception as exc:
        queue.put({"success": False, "error": str(exc)})


def run_analysis_direct(input_path: str, output_path: str, params: dict) -> None:
    from wenxing2 import run_analysis

    def ask_yes_no(title, message):
        if "地下水位" in message or "持力层位置" in title:
            return params.get("water_level_above", True)
        if "粉土" in title and "黏粒含量" in message:
            return params.get("silt_clay_content_ge_10", True)
        if "粉质黏土" in title or "粉质粘土" in title:
            return params.get("silty_clay_e_il_ge_085", True)
        return True

    run_analysis(None, input_path, output_path, ask_yes_no)


def execute_analysis(task_id: str, input_path: str, output_path: str, params: dict, timeout_seconds: int = 300) -> tuple[bool, str | None]:
    queue = multiprocessing.Queue()
    process = multiprocessing.Process(
        target=_run_analysis_in_process,
        args=(input_path, output_path, params, queue),
        daemon=True,
    )
    process.start()
    logger.info("任务 %s 已启动，进程 PID: %s", task_id, process.pid)
    process.join(timeout=timeout_seconds)

    if process.is_alive():
        process.terminate()
        process.join(timeout=5)
        if process.is_alive():
            process.kill()
            process.join(timeout=2)
        return False, f"分析超时（超过 {timeout_seconds} 秒），请检查文件是否过大或格式是否正确"

    try:
        result = queue.get_nowait()
    except Exception:
        if process.exitcode and process.exitcode != 0:
            return False, f"分析进程异常退出（退出码: {process.exitcode}）"
        return False, "分析进程异常退出"

    return result["success"], result["error"]


def process_analysis(task_id: str, input_path: str, output_path: str, params: dict | None = None) -> None:
    try:
        tasks[task_id]["status"] = "processing"
        tasks[task_id]["message"] = "正在分析中..."
        run_analysis_direct(input_path, output_path, params or {})
        tasks[task_id]["status"] = "completed"
        tasks[task_id]["message"] = "分析完成"
        tasks[task_id]["output_file"] = output_path
    except Exception as exc:
        tasks[task_id]["status"] = "failed"
        tasks[task_id]["message"] = f"分析失败: {exc}"
        cleanup_files(tasks[task_id])
        logger.exception("任务 %s 处理失败", task_id)


@app.exception_handler(Exception)
async def unhandled_exception_handler(request: Request, exc: Exception):
    logger.exception("请求异常: %s %s", request.method, request.url.path)
    return JSONResponse(status_code=500, content={"detail": "服务器内部异常，请稍后重试"})


def get_frontend_html() -> str:
    static_index_path = PROJECT_ROOT / "api" / "static" / "index.html"
    try:
        return static_index_path.read_text(encoding="utf-8")
    except Exception:
        logger.warning("读取正式前端页面失败，回退到内置页面: %s", static_index_path, exc_info=True)
        return INDEX_HTML


@app.get("/", response_class=HTMLResponse)
async def root():
    return HTMLResponse(get_frontend_html())


@app.get("/static/index.html", response_class=HTMLResponse)
async def static_index():
    return HTMLResponse(get_frontend_html())


@app.post("/api/precheck")
@limiter.limit("120/minute")
async def precheck_excel(request: Request, file: UploadFile = File(...)):
    if not is_allowed_excel_file(file.filename or ""):
        raise HTTPException(status_code=400, detail="只支持 Excel 文件 (.xlsx, .xls)")
    try:
        content = await file.read()
        workbook = load_workbook(filename=BytesIO(content), data_only=True)
        worksheet = workbook["1.6地层信息"] if "1.6地层信息" in workbook.sheetnames else None
        if worksheet is None:
            for name in workbook.sheetnames:
                if "地层" in name:
                    worksheet = workbook[name]
                    break
        has_silt = False
        has_silty_clay = False
        if worksheet is not None:
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue
                soil_name = ""
                if len(row) > 7 and row[7]:
                    soil_name = str(row[7]).strip().replace(" ", "").replace("\u3000", "")
                if "粉土" in soil_name:
                    has_silt = True
                if "粉质黏土" in soil_name or "粉质粘土" in soil_name:
                    has_silty_clay = True
                if has_silt and has_silty_clay:
                    break
        return {"has_silt": has_silt, "has_silty_clay": has_silty_clay}
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"预检失败: {exc}") from exc


@app.post("/api/upload")
@limiter.limit("10/minute")
async def upload_file(
    request: Request,
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    water_level_above: str = "true",
    silt_clay_content_ge_10: str = "true",
    silty_clay_e_il_ge_085: str = "true",
):
    cleanup_old_tasks()
    active_count = sum(1 for task in tasks.values() if task.get("status") in {"pending", "processing"})
    if active_count >= MAX_ACTIVE_TASKS:
        raise HTTPException(status_code=503, detail="当前分析任务过多，请稍后再试")

    original_filename = file.filename or ""
    if not is_allowed_excel_file(original_filename):
        raise HTTPException(status_code=400, detail="只支持 Excel 文件 (.xlsx, .xls)")

    try:
        file.file.seek(0, 2)
        file_size = file.file.tell()
        file.file.seek(0)
        max_size_bytes = MAX_FILE_SIZE_MB * 1024 * 1024
        if file_size > max_size_bytes:
            raise HTTPException(status_code=400, detail=f"文件过大，最大支持 {MAX_FILE_SIZE_MB}MB，当前文件 {file_size / 1024 / 1024:.1f}MB")

        params = {
            "water_level_above": water_level_above.lower() == "true",
            "silt_clay_content_ge_10": silt_clay_content_ge_10.lower() == "true",
            "silty_clay_e_il_ge_085": silty_clay_e_il_ge_085.lower() == "true",
        }
        logger.info("分析参数: %s", params)

        task_id = str(uuid.uuid4())[:8]
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        input_path = UPLOAD_DIR / f"{task_id}_{timestamp}_{sanitize_upload_filename(original_filename)}"
        output_path = OUTPUT_DIR / f"{task_id}_{timestamp}_分析报告.docx"

        with input_path.open("wb") as output_stream:
            shutil.copyfileobj(file.file, output_stream)

        tasks[task_id] = {
            "task_id": task_id,
            "status": "pending",
            "message": "任务已创建，等待处理",
            "output_file": "",
            "created_at": datetime.now().isoformat(),
            "input_file": str(input_path),
            "output_path": str(output_path),
            "params": params,
        }
        background_tasks.add_task(process_analysis, task_id, str(input_path), str(output_path), params)
        record_usage("analysis")
        return {"task_id": task_id, "message": "文件上传成功，开始分析"}
    except HTTPException:
        raise
    except Exception:
        logger.exception("保存上传文件或创建任务时失败: %s", original_filename)
        raise HTTPException(status_code=500, detail="文件上传失败，请稍后重试")


@app.get("/api/status/{task_id}")
async def get_task_status(task_id: str):
    if task_id not in tasks:
        raise HTTPException(status_code=404, detail="任务不存在")
    return tasks[task_id]


@app.get("/api/download/{task_id}")
async def download_result(task_id: str):
    task = tasks.get(task_id)
    if task is None:
        raise HTTPException(status_code=404, detail="任务不存在")
    if task["status"] != "completed":
        raise HTTPException(status_code=400, detail="任务尚未完成")
    output_path = Path(task["output_file"])
    if not output_path.exists():
        raise HTTPException(status_code=404, detail="输出文件不存在")
    return FileResponse(
        path=output_path,
        filename=f"地基基础分析报告_{task_id}.docx",
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    )


@app.get("/api/template")
async def download_template():
    template_path = PROJECT_ROOT / "Template_File" / "template_file.xlsx"
    if not template_path.exists():
        raise HTTPException(status_code=404, detail="模板文件不存在")
    return FileResponse(
        path=template_path,
        filename="输入数据模板.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.get("/api/health")
@limiter.limit("60/minute")
async def health_check(request: Request):
    return {
        "status": "ok",
        "message": "服务运行正常",
        "active_tasks": sum(1 for task in tasks.values() if task.get("status") in {"pending", "processing"}),
        "total_tasks": len(tasks),
        "runtime_root": str(RUNTIME_ROOT),
    }


@app.get("/api/stats/today")
@limiter.limit("60/minute")
async def api_today_stats(request: Request):
    return get_today_stats()


@app.get("/api/stats/summary")
@limiter.limit("30/minute")
async def api_stats_summary(request: Request, days: int = 7):
    days = max(1, min(days, 90))
    return get_stats_summary(days)


cleanup_thread = threading.Thread(target=scheduled_cleanup, daemon=True)
cleanup_thread.start()
logger.info("运行时目录: %s", RUNTIME_ROOT)
